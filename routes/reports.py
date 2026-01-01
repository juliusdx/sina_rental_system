
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from models import db, Invoice, InvoiceLineItem, Tenant
from datetime import datetime, date
from flask_login import login_required
from routes.auth import role_required
from utils import log_audit
import io
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    openpyxl = None

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/sst_preparation', methods=['GET', 'POST'])
@login_required
@role_required('admin', 'accounts')
def sst_preparation():
    if request.method == 'GET':
        return render_template('reports/sst_preparation.html')
    
    # Handle Report Generation
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid Booking Date', 'error')
        return redirect(url_for('reports.sst_preparation'))

    if not openpyxl:
        flash('Server missing openpyxl library', 'error')
        return redirect(url_for('billing.dashboard'))

    # Fetch Invoices
    # Filter by Issue Date? Or Receipt Date? 
    # SST is usually accrual basis (Invoice Date) for most businesses, 
    # but payment basis for some. MySST guide usually implies accrual (Invoice Issued).
    # We will use Invoice Issue Date.
    
    invoices = Invoice.query.filter(
        Invoice.issue_date >= start_date,
        Invoice.issue_date <= end_date,
        Invoice.status != 'void' # Exclude void
    ).order_by(Invoice.issue_date).all()
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "SST Return Draft"
    
    # Headers matching MySST Requirement Concepts
    headers = [
        'Invoice Date', 'Invoice No', 'Tenant Name', 'SST Reg No', 
        'Description', 'Taxable Service Value (Field 10)', 
        'SST 8% Charged (Field 12?)', 'Total Invoice Amount'
    ]
    
    # Style Headers
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    total_taxable_value = 0.0
    total_tax_charged = 0.0
    
    for inv in invoices:
        # We need to calculate how much of this invoice was taxable
        # We look for 'sst' line items to confirm it is a taxable invoice
        sst_lines = [li for li in inv.line_items if li.item_type == 'sst']
        
        if not sst_lines:
            continue # Skip non-taxable invoices? 
            # Or should we list them as "Exempt"? 
            # User wants "Taxable units", so likely only those we charged SST.
            
        tax_amount = sum(li.amount for li in sst_lines)
        
        # Calculate Taxable Value (The Base)
        # Assuming Tax = Base * 0.08  => Base = Tax / 0.08
        # OR Sum of all other line items? 
        # Safer to reverse calculate from Tax Amount to avoid floating point issues relative to what was charged.
        # But if we have mix of taxable and non-taxable items on one invoice, this is tricky.
        # Current logic: If we charge SST, we charge it on everything (Rent). 
        # Let's assume Base = Tax / 0.08
        
        taxable_value = tax_amount / 0.08
        
        row = [
            inv.issue_date,
            f"#{inv.id}",
            inv.tenant.name,
            inv.tenant.sst_registration_number or '-',
            inv.description,
            taxable_value,
            tax_amount,
            inv.total_amount
        ]
        ws.append(row)
        
        total_taxable_value += taxable_value
        total_tax_charged += tax_amount
        
    # Summary Row
    ws.append([])
    ws.append(['TOTAL', '', '', '', '', total_taxable_value, total_tax_charged])
    ws['F' + str(ws.max_row)].font = Font(bold=True)
    ws['G' + str(ws.max_row)].font = Font(bold=True)

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"SST_Report_{start_date_str}_to_{end_date_str}.xlsx"
    
    log_audit('REPORT', 'System', 0, f"Generated SST Report: {start_date_str} to {end_date_str}")
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route('/sst_exemptions', methods=['GET'])
@login_required
@role_required('admin', 'accounts')
def list_sst_exemptions():
    from models import SSTExemption
    exemptions = SSTExemption.query.join(Tenant).order_by(SSTExemption.start_date.desc()).all()
    return render_template('reports/sst_exemptions.html', exemptions=exemptions)
